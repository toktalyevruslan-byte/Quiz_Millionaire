from __future__ import annotations

import csv
import json
import os
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence


def _format_time_mmss(time_seconds: Any) -> str:
    try:
        total = int(time_seconds or 0)
    except Exception:
        total = 0
    minutes = total // 60
    seconds = total % 60
    return f"{minutes}:{seconds:02d}"


def _format_rub(score: Any, symbol: str = "₽") -> str:
    try:
        value = int(score or 0)
    except Exception:
        value = 0
    # Use spaces as thousands separators to match UI style
    return f"{value:,} {symbol}".replace(",", " ")


@dataclass(frozen=True)
class ExportResult:
    file_path: str
    exports_dir: str


class DataExporter:
    """
    Export leaderboard (Top-10) to JSON/CSV/Excel/PDF into exports/ with timestamped filenames.

    Expected record dict keys (from DatabaseManager.get_top_scores):
      - name, score, rank (series), time_seconds, date
    """

    def __init__(self, base_dir: Optional[str] = None) -> None:
        self.base_dir = base_dir or os.path.dirname(os.path.abspath(__file__))
        self.exports_dir = os.path.join(self.base_dir, "exports")
        os.makedirs(self.exports_dir, exist_ok=True)

    def _timestamp(self) -> str:
        return datetime.now().strftime("%Y%m%d_%H%M%S")

    def _normalize_rows(self, data: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for i, rec in enumerate(data, 1):
            name = str(rec.get("name", "")).strip()
            try:
                series = int(rec.get("rank", 0) or 0)
            except Exception:
                series = 0
            time_seconds = rec.get("time_seconds", 0) or 0
            score = rec.get("score", 0) or 0

            rows.append(
                {
                    "№": i,
                    "Имя": name,
                    "Серия": series,
                    "Время": _format_time_mmss(time_seconds),
                    "Выигрыш": _format_rub(score),
                }
            )
        return rows

    def export_json(self, leaderboard: Sequence[Dict[str, Any]]) -> ExportResult:
        rows = self._normalize_rows(leaderboard)
        filename = f"leaderboard_export_{self._timestamp()}.json"
        path = os.path.join(self.exports_dir, filename)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, indent=2)
        return ExportResult(file_path=path, exports_dir=self.exports_dir)

    def export_csv(self, leaderboard: Sequence[Dict[str, Any]]) -> ExportResult:
        rows = self._normalize_rows(leaderboard)
        filename = f"leaderboard_export_{self._timestamp()}.csv"
        path = os.path.join(self.exports_dir, filename)
        fieldnames = ["№", "Имя", "Серия", "Время", "Выигрыш"]
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for r in rows:
                writer.writerow({k: r.get(k, "") for k in fieldnames})
        return ExportResult(file_path=path, exports_dir=self.exports_dir)

    def export_excel(self, leaderboard: Sequence[Dict[str, Any]]) -> ExportResult:
        try:
            import pandas as pd  # type: ignore
        except Exception as e:
            raise RuntimeError("Excel export requires pandas + openpyxl. Install: pip install pandas openpyxl") from e

        filename = f"leaderboard_export_{self._timestamp()}.xlsx"
        path = os.path.join(self.exports_dir, filename)

        rows = self._normalize_rows(leaderboard)
        df = pd.DataFrame(rows, columns=["№", "Имя", "Серия", "Время", "Выигрыш"])

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Leaderboard")

            # Light formatting (optional)
            try:
                ws = writer.sheets["Leaderboard"]
                from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore

                header_fill = PatternFill("solid", fgColor="10233F")
                header_font = Font(bold=True, color="FFD700")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                widths = {"A": 6, "B": 20, "C": 10, "D": 10, "E": 16}
                for col, w in widths.items():
                    ws.column_dimensions[col].width = w
            except Exception:
                pass

        return ExportResult(file_path=path, exports_dir=self.exports_dir)

    def export_pdf(self, leaderboard: Sequence[Dict[str, Any]]) -> ExportResult:
        try:
            from reportlab.lib import colors  # type: ignore
            from reportlab.lib.pagesizes import A4  # type: ignore
            from reportlab.lib.styles import getSampleStyleSheet  # type: ignore
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle  # type: ignore
        except Exception as e:
            raise RuntimeError("PDF export requires reportlab. Install: pip install reportlab") from e

        filename = f"leaderboard_export_{self._timestamp()}.pdf"
        path = os.path.join(self.exports_dir, filename)

        rows = self._normalize_rows(leaderboard)
        table_data: List[List[Any]] = [["№", "Имя", "Серия", "Время", "Выигрыш"]]
        for r in rows:
            table_data.append([r["№"], r["Имя"], r["Серия"], r["Время"], r["Выигрыш"]])

        doc = SimpleDocTemplate(path, pagesize=A4, topMargin=36, bottomMargin=36, leftMargin=36, rightMargin=36)
        styles = getSampleStyleSheet()
        story: List[Any] = []
        story.append(Paragraph("Таблица рекордов (Топ-10)", styles["Title"]))
        story.append(Spacer(1, 12))

        table = Table(table_data, repeatRows=1, colWidths=[30, 170, 60, 60, 90])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10233F")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#FFD700")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#2a3b66")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F3F6FF"), colors.white]),
                    ("ALIGN", (0, 1), (0, -1), "CENTER"),
                    ("ALIGN", (2, 1), (3, -1), "CENTER"),
                    ("ALIGN", (4, 1), (4, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )

        story.append(table)
        doc.build(story)

        return ExportResult(file_path=path, exports_dir=self.exports_dir)

